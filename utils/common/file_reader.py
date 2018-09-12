#-*- coing=utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from config.globalconfig import TEST_DATA_PATH,BASE_PATH,TEST_APP_PATH
import chardet
import os.path
import configparser
from configparser import NoOptionError, NoSectionError
from utils.common.log import *
import codecs

filename= os.path.basename(os.path.realpath(__file__)).split('.')[0]
logger = Logger(filename).get_logger()
config_path = os.path.join(BASE_PATH, 'config', 'config.ini')

class ExcelReader(object):

    def __init__(self, excel, sheet=0, title_line=True,column=None,row=None):
        """
        :param excel: excel path,type=str
        :param sheet: sheet name or sheet sequence number,type = str or int
        :param title_line: title_line = True,return dic,title_line = False,return list
        :param column: Non essential parameter,type = int or str,int express column number，str express column field name of first line
        :param row: Non essential parameter,type = int,express row number
        """
        if os.path.exists(excel):
           self.excel = excel
        else:
            raise FileNotFoundError('Excel file does not exist ！')
        self.sheet = sheet
        self.title_line = title_line
        self.column = column
        self.row = row
        self._data = []
        self._cell_data = str()
        self.workbook = load_workbook(self.excel)
        if type(self.sheet) not in [int, str]:
            raise TypeError('Please pass in <type int> or <type str>, not {0}'.format(type(self.sheet)))
        elif type(self.sheet) == int:
            sheetnames = self.workbook.sheetnames
            self.ws = self.workbook[sheetnames[self.sheet]]
        else:
            self.ws = self.workbook[self.sheet]

    @property
    def sheet_data(self):
        """
            Read the contents of the excel file, Return list.
            example:
            The contents of the excel file：
            | A  | B  | C  |
            | A1 | B1 | C1 |
            | A2 | B2 | C2 |
            print(ExcelReader(excel, title_line=True).data)，output：
            [{A: A1, B: B1, C:C1}, {A:A2, B:B2, C:C2}]
            print(ExcelReader(excel, title_line=False).data)，output：
            [[A,B,C], [A1,B1,C1], [A2,B2,C2]]
            You can specify sheet through index or name:
            example:
            ExcelReader(excel, sheet=2)
            ExcelReader(excel, sheet='testdata')
        """
        if not self._data:
            if self.title_line:
                for row in self.ws.iter_rows(min_row=1, max_row=1, ):
                    title = [cell.value for cell in row]# 首行为title

                for i in range(1, self.ws.max_row):
                    every_row_cell_list = list(self.ws.rows)[i]
                    every_row_value = []
                    for cell in every_row_cell_list:
                        every_row_value.append(cell.value)
                    self._data.append(dict(zip(title, every_row_value)))
            else:
                for row in self.ws.rows:
                    self._data.append([cell.value for cell in row])
        return self._data

    def column_letter(self):
        """Column number converted to letter form."""
        if isinstance(self.column, (int)):
            return get_column_letter(self.column)  # 数字转换为字母
        elif isinstance(self.column, (str)):
            for row in self.ws.iter_rows(min_row=1, max_row=1, ):
                columnFieldName = [cell.value for cell in row]
            if columnFieldName.count(self.column) == 1:  # 相同字段名存在的个数
                for index, cell in enumerate(columnFieldName):
                    if cell == self.column:
                        return get_column_letter(index + 1)
            elif columnFieldName.count(self.column) == 0:
                raise ValueError("Current sheet page '%s' no field name '%s' exists!" % (self.ws.title, self.column))
            else:
                raise ValueError("Current sheet page'%s'include duplicate field name '%s'" % (self.ws.title, self.column))
        else:
            raise TypeError("The column parameter type should be int or str, not '%s'" % (self.column.__class__))
    @property
    def cell_data(self):
        if not self._cell_data:
            if not isinstance(self.row, (int)):
                raise TypeError("The row parameter type should be int instead of '%s'!" % (self.row.__class__))
            column = self.column_letter()
            return self.ws[column + str(self.row)].value


class ExcelWritter(ExcelReader):
    def __init__(self, excel, sheet=0, column=None,row=None,value=''):
        ExcelReader.__init__(self,excel,sheet=0)
        self.value = value
        self.column =column
        self.row = row

    def write_cell(self):
        """Write cell value of Excel file"""
        if not isinstance(self.row, (int)):
            raise TypeError("The row parameter type should be int instead of '%s'!" % (self.row.__class__))
        if not isinstance(self.value, (str)):
            raise TypeError("The value parameter type should be str instead of '%s'!" % (self.value.__class__))

        column = self.column_letter()
        self.ws[column + str(self.row)].value = self.value
        self.workbook.save(self.excel)


def getFileCoding(filepath: str) -> dict:
    """
    func：Get the encoding format of the file
    :param filepath: type=str
    :return: {'confidence': 1.0, 'encoding': 'UTF-8'}, type=dict
    """
    with open(file=filepath, mode="rb") as f:
        s = f.read()
    return chardet.detect(s)

class ConfiginiReader(object):

    def __init__(self,configPath):
        """
        func:read congig.ini file
        :param configpath: config file path,type=str
        :return congfig content
        """
        self.configPath = configPath

    def config_data(self):
        config = configparser.ConfigParser()

        with open(file=self.configPath, mode="rb") as f:  # 2018-6-10 BEGIN 判断文件的编码类型 如果是UTF-8 BOM格式，则将其转化为无BOM格式
            s = f.read()
        if s.startswith(codecs.BOM_UTF8):  # 带BOM的文件是以 b'\xef\xbb\xbf' 开头
            s = s[len(codecs.BOM_UTF8):]  # 截取 b'\xef\xbb\xbf' 到文件结尾
            with open(file=self.configPath, mode="wb") as f:  # 保存为无BOM格式
                f.write(s)
        coding = getFileCoding(filepath=self.configPath).get("encoding")
        config.read(filenames=self.configPath, encoding=coding)  # 2018-6-10 END 判断文件的编码类型 如果是UTF-8 BOM格式，则将其转化为无BOM格式
        return config

    @property
    def TestAppPath(self):
        try:
            testAppPath = self.config_data().get(section='TEST_APP_PATH',option='test_app_path')
        except(NoOptionError, NoSectionError):
            testAppPath = TEST_APP_PATH
            logger.warning(
                'The TEST_APP_PATH.test_app_path option is not configured in the config.ini file, and the default value "%s" will be used.',
                testAppPath)
        return testAppPath

    @property
    def timeOut(self):
        try:
            timeout = self.config_data().getint(section="TEST", option="timeOut")
        except (NoOptionError, NoSectionError):
            timeout = 10
            logger.warning(
                'The TEST.timeOut option is not configured in the config.ini file, and the default value "%s" will be used.', timeout)
        return timeout

    @property
    def headLess(self):
        try:
            headless = self.config_data().getboolean(section="TEST", option="headLess")
        except (NoOptionError, NoSectionError):
            headless = False
            logger.warning(
                'The TEST.headLess option is not configured in the config.ini file, and the default value "%s" will be used.',
                headless)
        return headless

    @property
    def sendEmail(self):
        try:
            sendemail = self.config_data().getboolean(section="EMAIL", option="sendEmail")
        except (NoOptionError, NoSectionError):
            sendemail = False
            logger.warning(
                'The EMAIL.sendEmail option is not configured in the config.ini file, and the default value "%s" will be used.',
                sendemail)
        return sendemail

    @property
    def emailSubject(self):
        try:
            emailTitle = self.config_data().get(section="EMAIL", option="title")
        except (NoOptionError, NoSectionError):
            emailTitle = "WonderStitch自动化测试报告邮件"
            logger.warning(
                'The EMAIL.title option is not configured in the config.ini file, and the default value "%s" will be used.',
                emailTitle)
        return emailTitle

    @property
    def passwd(self):
        try:
            password = self.config_data().get(section="EMAIL", option="SMTPLoginPasswd")
            return password
        except (NoOptionError, NoSectionError):
            logger.warning('The EMAIL.SMTPLoginPasswd option is not configured in the config.ini file')

    @property
    def sender(self):
        try:
            emailSender= self.config_data().get(section="EMAIL", option="senderAddress")
            return emailSender
        except (NoOptionError, NoSectionError):
            logger.warning('The EMAIL.senderAddress option is not configured in the config.ini file')

    @property
    def receivers(self):
        try:
            address = self.config_data().items(section="receiversAddress")  # [(,), (,)...]
            return list(dict(address).values())
        except (NoOptionError, NoSectionError):
            logger.warning('The receiversAddress section is not configured in the config.ini file')

    @property
    def SMTPServer(self):
        try:
            swapServer = self.config_data().get(section="EMAIL", option="server")
            return swapServer
        except (NoOptionError, NoSectionError):
            logger.warning('The EMAIL.server option is not configured in the config.ini file')

    @property
    def SMTPPort(self):
        try:
            swapPort = self.config_data().getint(section="EMAIL", option="SMTPPort")
            return swapPort
        except (NoOptionError, NoSectionError):
            logger.warning('The EMAIL.SMTPPort option is not configured in the config.ini file')

    @property
    def stitchOutputPath(self):
        try:
            videoOutputPath = self.config_data().get(section="TestResultDataPath", option="stitchOutputPath")
            return videoOutputPath
        except (NoOptionError, NoSectionError):
            logger.warning('The TestResultDataPath.stitchOutputPath option is not configured in the config.ini file')

class Config(object):
    configReader = ConfiginiReader(configPath = config_path)
    #待测软件路径
    testAppPath = configReader.TestAppPath
    # 配置隐形等待时间
    timeOut = configReader.timeOut
    # 无头模式运行
    headLess = configReader.headLess
    # Email
    # 是否发送邮件
    sendEmail = configReader.sendEmail
    # SMTP服务器登录名
    emailSender = configReader.sender
    # SMTP服务器登录密码
    emialPasswd =configReader.passwd
    # 收件地址
    receivers = configReader.receivers
    # SMTP服务器地址
    SMTPServer= configReader.SMTPServer
    # SMTP服务器地址端口
    SMTPPort = configReader.SMTPPort
    # 邮件主题
    emailTitle = configReader.emailSubject
    #输出拼接视频和拼接图片地址
    stitchOutputPath = configReader.stitchOutputPath


# if __name__ == '__main__':
#     test_data = os.path.join(TEST_DATA_PATH,'TestCase.xlsx')
#     reader = ExcelReader(test_data,sheet= 0,title_line=True,column='stitch_type',row=4)
#     print(type(reader.sheet_data[0]['测试参数']))
#     print(eval(reader.sheet_data[1]['测试参数']))
    # list = ['H264','H265']
    # if reader.sheet_data[1][ 'codec_type'].upper() not in  list:
    #     print('Y')
    #
    # print(reader.cell_data)
    # writter = ExcelWritter(test_data,sheet= 0,column='stitch_type',row=5,value='h264')
    # writter.write_cell()
    #
    # print(config_path)
    #config = Config()

